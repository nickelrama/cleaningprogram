from tkinter import *

import xlrd
import xlwt
import getpass
from operator import itemgetter
from datetime import date
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
import random
import os
from dateutil.parser import parse
import collections


class Application(object):

    def __init__(self, event=None):

        self.root = Tk()

        self.root.configure(bg="darkorchid1", padx=10, pady=10)
        self.root.title("WELCOME TO NCP (NICKELRAMA CLEANING PROGRAM)")

        self.nickelrama = Label(self.root, text="NCP - NICKELRAMA CLEANING PROGRAM", bg="lightgrey", fg="black")
        self.nickelrama.pack()

        ##################

        userframe = Frame(self.root)
        userframe.configure(bg="magenta4")
        userframe.pack()

        self.enternamelabel = Label(userframe, text="ENTER NAME EXACTLY AS IT APPEARS ON SCHEDULE:", bg="purple", fg="white", width=50)
        self.enternamelabel.pack()

        self.entryuser = Entry(userframe, width=25)
        self.entryuser.configure(fg= "white",bg="grey20")
        self.entryuser.pack()
        self.root.bind("<Return>", self.submit)
#        return self.entryuser

        ####################

        passwordframe = Frame(self.root)
        passwordframe.configure(bg="magenta4")
        passwordframe.pack()

        self.enterpasswordlabel = Label(passwordframe, text="ENTER PASSWORD", bg="purple", fg="white", width=50)
        self.enterpasswordlabel.pack()

        self.entrypassword = Entry(passwordframe, width=25)
        self.entrypassword.configure(fg="grey20", bg="grey20")
        self.entrypassword.pack()
#        return self.entrypassword

        self.submitbutton = Button(self.root, text="SUBMIT", highlightbackground="green", width=48, padx=10, pady=10, command=self.submit)
        self.submitbutton.pack()




#############DEFINITIONS###################
        self.username = "Cody"
        self.password = "monticello"
        self.tasksbutton=None
        self.loginbutton = None
        self.incorrectbutton = None
        self.attempts = 0

#        self.cell = None
#        self.cell.value = return
#        self.val = None
#        self.c = self.cell.coordinate
#        self.offset3 = self.cell.offset(0, -1).value
#        self.datedata = self.cell.offset(0, -3).value
#        self.offset2 = self.cell.offset(0, -2).value
#        self.offset1 = self.cell.offset(0, -3).value
#        self.tasks = Label("{0} | {1} | {2} | {3}".format(self.offset1, self.offset2, self.offset3, val))
###########################################



    def submit(self, event=None):
        username = self.entryuser.get()
        password = self.entrypassword.get()
        if username == self.username and password == self.password:
                self.loginbutton = Button(self.root, text='LOGIN', highlightbackground="green", width=28, command=self.login)
                self.loginbutton.pack()
                self.root.bind("<Return>", self.login)
                self.submitbutton.config(state="disabled")

        else:
                self.incorrectbutton = Button(self.root, text="INCORRECT- CLICK/ENTER TO DIMISS THIS MESSAGE", highlightbackground="red", width=48, padx=10, pady=10, command=self.incorrect)
                self.incorrectbutton.pack()
                self.root.bind("<Return>", self.incorrect)
                self.submitbutton.config(state="disabled")

    def incorrect(self, event=None):
        self.attempts += 1
        if (self.attempts >2):
            self.root.destroy()


        else:
            self.root.bind("<Return>", self.submit)
            self.submitbutton.config(state="normal", padx=10, pady=10)
            self.incorrectbutton.destroy()
            self.incorrectbutton = None
            self.entrypassword.delete(0, END)
            self.entryuser.delete(0, END)

#####################################################

    def login(self, event=None):

        username = self.entryuser.get()

        self.employee = Label(self.root, text= "Welcome %s" % username, bg="lightgrey", fg="black")
        self.employee.pack()

        self.tasksbutton = Button(self.root, text= "Click here in order to view tasks currently in progress.", command = self.main)
        self.tasksbutton.pack()
        self.root.bind("<Return>", self.main)
        self.loginbutton.config(state="disabled")


    def main(self, event=None):

        self.currenttasks = Label(self.root, text='These are the tasks currently in progress: ')
        self.currenttasks.pack()

        self.val = self.cell.value
        self.c = self.cell.coordinate
        self.offset3 = self.cell.offset(0, -1).value
        self.datedata = self.cell.offset(0, -3).value
        self.offset2 = self.cell.offset(0, -2).value
        self.offset1 = self.cell.offset(0, -3).value
        self.tasks = Label("{0} | {1} | {2} | {3}".format(self.offset1, self.offset2, self.offset3, val))

        self.wb = load_workbook('Cleaning.xlsx')
        self.ws = self.wb.get_sheet_by_name("Cleaning")
        self.x = Label(text='In progress')

        for row in ws.rows:
            for self.cell in row:
                if self.cell.value == self.x:
                    self.val = cell.value
                    self.c = cell.coordinate
                    self.offset3 = self.cell.offset(0, -1).value
                    self.datedata = self.cell.offset(0, -3).value
                    self.offset2 = self.cell.offset(0, -2).value
                    self.offset1 = self.cell.offset(0, -3).value


    def userlvl(self, event=None):
        username = Cody
        if username == 1:
            trainee= Text('Hello Trainee', command = self.traineeselections)
        elif username == 2:
            print('Hello Attendent')
            attendantselections(username)
        elif username == 3:
            print('Hello Staff Attendent')
            staffselections(username)
        elif username == 4:
            print('Senior Staff')
            seniorselections(username)
        elif username == Cody:
            manager_options(username)
































        ###########****************************************************************************################
        def login():
            ### Manager List
            book_manager = xlrd.open_workbook('masterlist.xlsx')
            sheet = book_manager.sheet_by_name('Managers')
            data_manager = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            data_manager = [item for sublist in data_manager for item in sublist]
            manager = map(lambda it: it.strip(), data_manager)

            ###Trainee list
            book_trainee = xlrd.open_workbook('masterlist.xlsx')
            sheet = book_trainee.sheet_by_name('Trainee')
            data_trainee = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            data_trainee = [item for sublist in data_trainee for item in sublist]
            trainee = map(lambda it: it.strip(), data_trainee)

            ####Attendant list
            book_attendant = xlrd.open_workbook('masterlist.xlsx')
            sheet = book_attendant.sheet_by_name('Attendant')
            data_attendant = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            data_attendant = [item for sublist in data_attendant for item in sublist]
            attendant = map(lambda it: it.strip(), data_attendant)

            ####Staff Attendant list
            book_staff_attendant = xlrd.open_workbook('masterlist.xlsx')
            sheet = book_staff_attendant.sheet_by_name('Staff Attendant')
            data_staff_attendant = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            data_staff_attendant = [item for sublist in data_staff_attendant for item in sublist]
            staff_attendant = map(lambda it: it.strip(), data_staff_attendant)

            ####Senior Staff list
            book_senior_staff = xlrd.open_workbook('masterlist.xlsx')
            sheet = book_senior_staff.sheet_by_name('Senior Staff')
            data_senior_staff = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            data_senior_staff = [item for sublist in data_senior_staff for item in sublist]
            senior_staff = map(lambda it: it.strip(), data_senior_staff)

            ####
            username = str(input('Please enter your name as it appears on the schedule:'))
            if username in trainee:
                os.system('cls')
                print('Hello it worked', username, 'you are a Trainee')
                return 1, username
            elif username in attendant:
                os.system('cls')
                print('Hello it worked', username, 'you are an Attendant')
                return 2, username
            elif username in staff_attendant:
                os.system('cls')
                print('Hello it worked', username, 'you are a Staff Attendant')
                return 3, username
            elif username in senior_staff:
                os.system('cls')
                print('Hello it worked', username, 'you are Senior Staff')
                return 4, username
            elif username in manager:
                os.system('cls')
                p = getpass.getpass(prompt='Please enter the password: ')
                if p == "monticello":
                    os.system('cls')
                    print('Password Accepted')
                    print('Hello', username, 'you are a part of the Manager group')
                    return 5, username

        #######################
        def traineeselections(username):
            while True:
                try:
                    count = 1
                    os.system('cls')

                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    x = 'In progress'
                    in_progress = []
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == x:
                                val = cell.value
                                offset3 = cell.offset(0, -1).value
                                in_progress.append(offset3)

                                ###
                    print('Please select a task from the available options')
                    book_trainee_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_trainee_tasks.sheet_by_name('trainee_tasks')
                    data_trainee_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                          range(sheet.nrows)]
                    data_trainee_tasks1 = [item for sublist in data_trainee_tasks for item in sublist]
                    data_trainee_tasks = [item for item in data_trainee_tasks1 if item not in in_progress]
                    j = (len(data_trainee_tasks))
                    length = int(j)
                    length2 = length
                    length2 = length2 + 1
                    num = range(1, length2)
                    numlist = map(str, num)
                    for row in zip(numlist, data_trainee_tasks):
                        print('.'.join(row))
                    x = int(input('Enter a selection 1 - ' + str(length) + ': '))
                    x = x - 1
                    selection = itemgetter(x)(data_trainee_tasks)
                    print(selection)
                    wb = load_workbook('Cleaning.xlsx')
                    print('Loading Workbook')
                    ws = wb.get_sheet_by_name("Cleaning")
                    print('Getting Sheet')
                    today = date.today()
                    stringdate = today.strftime("%Y-%m-%d %H:%M:%S")
                    dt = parse(stringdate)
                    fixed = dt.strftime('%m/%d/%Y')
                    ws.append([fixed, username, selection, 'In progress'])
                    print('Writing to sheet')
                    ws.column_dimensions['A'].width = 11
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 50
                    ws.column_dimensions['D'].width = 16
                    ws.column_dimensions['E'].width = 30
                    wb.save('Cleaning.xlsx')
                    print('Workbook saved')
                    print('       ')
                    print('       ')
                    count = count + 1
                    if count > 1:
                        input('Press Enter to return to the login screen')
                        break

                except:
                    print('Number is out of range of the list, please try again.')
                    input('Press enter to try again')
                    pass

        def attendantselections(username):
            while True:
                try:
                    count = 1
                    os.system('cls')
                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    x = 'In progress'
                    in_progress = []
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == x:
                                val = cell.value
                                offset3 = cell.offset(0, -1).value
                                in_progress.append(offset3)

                    print('Please select a task from the available options')
                    book_trainee_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_trainee_tasks.sheet_by_name('trainee_tasks')
                    data_trainee_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                          range(sheet.nrows)]
                    book_attendant_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_attendant_tasks.sheet_by_name('attendant_tasks')
                    data_attendant_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                            range(sheet.nrows)]
                    data_attendant_tasks_merged = data_trainee_tasks + data_attendant_tasks
                    data_attendant_tasks_merged1 = [item for sublist in data_attendant_tasks_merged for item in sublist]
                    data_attendant_tasks_merged = [item for item in data_attendant_tasks_merged1 if
                                                   item not in in_progress]

                    j = (len(data_attendant_tasks_merged))
                    length = int(j)
                    length2 = length
                    length2 = length2 + 1
                    num = range(1, length2)
                    numlist = map(str, num)
                    for row in zip(numlist, data_attendant_tasks_merged):
                        print('.'.join(row))
                    x = int(input('Enter a selection 1 - ' + str(length) + ': '))
                    x = x - 1
                    selection = itemgetter(x)(data_attendant_tasks_merged)
                    print(selection)
                    wb = load_workbook('Cleaning.xlsx')
                    print('Loading Workbook')
                    ws = wb.get_sheet_by_name("Cleaning")
                    print('Getting Sheet')
                    today = date.today()
                    stringdate = today.strftime("%Y-%m-%d %H:%M:%S")
                    dt = parse(stringdate)
                    fixed = dt.strftime('%m/%d/%Y')
                    ws.append([fixed, username, selection, 'In progress'])
                    print('Writing to sheet')
                    ws.column_dimensions['A'].width = 11
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 50
                    ws.column_dimensions['D'].width = 16
                    ws.column_dimensions['E'].width = 30
                    wb.save('Cleaning.xlsx')
                    print('Workbook saved')
                    print('       ')
                    print('       ')
                    count = count + 1
                    if count > 1:
                        input('Press Enter to return to the login screen')
                        break

                except:
                    print('Number is out of range of the list, please try again.')
                    input('Press enter to try again')
                    pass

        def staffselections(username):
            while True:
                try:
                    count = 1
                    os.system('cls')
                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    x = 'In progress'
                    in_progress = []
                    in_progress_1 = []
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == x:
                                val = cell.value
                                offset3 = cell.offset(0, -1).value
                                in_progress_1.append(offset3)

                    print('Please select a task from the available options')
                    book_trainee_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_trainee_tasks.sheet_by_name('trainee_tasks')
                    data_trainee_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                          range(sheet.nrows)]
                    book_attendant_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_attendant_tasks.sheet_by_name('attendant_tasks')
                    data_attendant_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                            range(sheet.nrows)]
                    book_staff_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_staff_tasks.sheet_by_name('staff_tasks')
                    data_staff_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                        range(sheet.nrows)]
                    data_staff_tasks1 = [item for sublist in data_staff_tasks for item in sublist]
                    data_staff_tasks_merged1 = data_trainee_tasks + data_attendant_tasks + data_staff_tasks
                    data_staff_tasks_merged12 = [item for sublist in data_staff_tasks_merged1 for item in sublist]
                    in_progress = [item for item in in_progress_1 if item not in data_staff_tasks1]
                    print(in_progress)

                    data_staff_tasks_merged = [item for item in data_staff_tasks_merged12 if item not in in_progress]
                    j = (len(data_staff_tasks_merged))
                    length = int(j)
                    length2 = length
                    length2 = length2 + 1
                    num = range(1, length2)
                    numlist = map(str, num)
                    for row in zip(numlist, data_staff_tasks_merged):
                        print('.'.join(row))
                    x = int(input('Enter a selection 1 - ' + str(length) + ': '))
                    x = x - 1
                    selection = itemgetter(x)(data_staff_tasks_merged)
                    print(selection)
                    wb = load_workbook('Cleaning.xlsx')
                    print('Loading Workbook')
                    ws = wb.get_sheet_by_name("Cleaning")
                    print('Getting Sheet')
                    today = date.today()
                    stringdate = today.strftime("%Y-%m-%d %H:%M:%S")
                    dt = parse(stringdate)
                    fixed = dt.strftime('%m/%d/%Y')
                    ws.append([fixed, username, selection, 'In progress'])
                    print('Writing to sheet')
                    ws.column_dimensions['A'].width = 11
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 50
                    ws.column_dimensions['D'].width = 16
                    ws.column_dimensions['E'].width = 30
                    wb.save('Cleaning.xlsx')
                    print('Workbook Saved')
                    print('       ')
                    print('       ')
                    count = count + 1
                    if count > 1:
                        input('Press Enter to return to the login screen')
                        break

                except:
                    print('Number is out of range of the list, please try again.')
                    input('Press enter to try again')
                    pass

        ##
        def seniorselections(username):
            while True:
                try:
                    count = 1
                    os.system('cls')
                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    x = 'In progress'
                    in_progress = []
                    in_progress_1 = []
                    in_progress_2 = []
                    in_progress_3 = []
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == x:
                                val = cell.value
                                offset3 = cell.offset(0, -1).value
                                in_progress_1.append(offset3)

                    print('Please select a task from the available options')
                    book_trainee_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_trainee_tasks.sheet_by_name('trainee_tasks')
                    data_trainee_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                          range(sheet.nrows)]
                    book_attendant_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_attendant_tasks.sheet_by_name('attendant_tasks')
                    data_attendant_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                            range(sheet.nrows)]

                    book_staff_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_staff_tasks.sheet_by_name('staff_tasks')
                    data_staff_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                        range(sheet.nrows)]
                    data_staff_tasks1 = [item for sublist in data_staff_tasks for item in sublist]
                    book_senior_tasks = xlrd.open_workbook('masterlist.xlsx')
                    sheet = book_senior_tasks.sheet_by_name('senior_tasks')
                    data_senior_tasks = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in
                                         range(sheet.nrows)]
                    data_senior_tasks1 = [item for sublist in data_senior_tasks for item in sublist]
                    data_senior_tasks_merged1 = data_trainee_tasks + data_attendant_tasks + data_staff_tasks + data_senior_tasks
                    in_progress_2 = [item for item in in_progress_1 if item not in data_staff_tasks1]
                    in_progress = [item for item in in_progress_2 if item not in data_senior_tasks1]
                    print(in_progress)
                    data_senior_tasks_merged12 = [item for sublist in data_senior_tasks_merged1 for item in sublist]
                    data_senior_tasks_merged = [item for item in data_senior_tasks_merged12 if item not in in_progress]
                    j = (len(data_senior_tasks_merged))
                    length = int(j)
                    length2 = length
                    length2 = length2 + 1
                    num = range(1, length2)
                    numlist = map(str, num)
                    for row in zip(numlist, data_senior_tasks_merged):
                        print('.'.join(row))
                    x = int(input('Enter a selection 1 - ' + str(length) + ': '))
                    x = x - 1
                    selection = itemgetter(x)(data_senior_tasks_merged)
                    print(selection)
                    wb = load_workbook('Cleaning.xlsx')
                    print('Loading Workbook')
                    ws = wb.get_sheet_by_name("Cleaning")
                    print('Getting Sheet')
                    today = date.today()
                    stringdate = today.strftime("%Y-%m-%d %H:%M:%S")
                    dt = parse(stringdate)
                    fixed = dt.strftime('%m/%d/%Y')
                    ws.append([fixed, username, selection, 'In progress'])
                    print('Writing to sheet')
                    ws.column_dimensions['A'].width = 11
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 50
                    ws.column_dimensions['D'].width = 16
                    ws.column_dimensions['E'].width = 30
                    wb.save('Cleaning.xlsx')
                    print('Workbook saved')
                    print('       ')
                    print('       ')
                    count = count + 1
                    if count > 1:
                        input('Press Enter to return to the login screen')
                        break
                except:
                    print('Number is out of range of the list, please try again.')
                    input('Press enter to try again')
                    pass

        def managerselection(username):
            while True:
                try:
                    os.system('cls')
                    print('These are the tasks that are currently in progress')
                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    x = 'In progress'
                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    check = 'Completed'
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == x:
                                val = cell.value
                                c = cell.coordinate
                                offset3 = cell.offset(0, -1).value
                                offset2 = cell.offset(0, -2).value
                                offset1 = cell.offset(0, -3).value
                                print('   ')
                                print("{0} | {1} | {2} | {3} | {4}".format(c, offset1, offset2, offset3, val))
                    test = input('Input Coordinates of cell to change, enter exit to return to the manager menu: ')
                    if test == 'exit':
                        manager_options(username)
                        break

                    test1 = test
                    length = (len(test1))
                    c = ws.cell(test)
                    for row in ws.rows:
                        for cell in row:
                            if c.value == check:
                                print('Cell has been marked completed')
                                input('Press Enter to continue')
                                managerselection(username)

                            else:
                                c.value = 'Completed'
                                c.offset(0, 1).value = username
                                ws.column_dimensions['A'].width = 11
                                ws.column_dimensions['B'].width = 30
                                ws.column_dimensions['C'].width = 50
                                ws.column_dimensions['D'].width = 16
                                ws.column_dimensions['E'].width = 30
                                wb.save('Cleaning.xlsx')
                                print('       ')


                except:
                    print('Invalid Input, please try again')
                    input('Press Enter to try again')
                    pass

        ###################
        def overwrite():
            wb = Workbook()
            print('Good')
            ws = wb.active
            ws.title = "Games to Clean"
            book = xlrd.open_workbook('masterlist.xlsx')
            sheet = book.sheet_by_name('Games')
            all_games = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
            all_games = [item for sublist in all_games for item in sublist]
            print(len(all_games))
            bookemp = xlrd.open_workbook('masterlist.xlsx')
            sheet = bookemp.sheet_by_name('Employees')
            employee_list = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            employee_list = [item for sublist in employee_list for item in sublist]
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws['A1'] = 'Employee'
            ws['B1'] = 'Game to Clean'
            ws['C1'] = 'Manager Initial'
            ws['D1'] = 'Date'
            num = (len(employee_list))
            games_to_clean = []
            count = 0
            while count < num:
                random.shuffle(all_games)
                x = all_games.pop()
                games_to_clean.append(x)
                count = count + 1
            for row in zip(employee_list, games_to_clean):
                ws.append(row)
            print(all_games)
            print(len(all_games))
            wb.save(filename='data.xlsx')
            print('DONE')
            return all_games

        def writememory(all_games):
            if isinstance(all_games, str):
                print('Reset needed')
            else:
                wb = Workbook()
                print('Good')
                ws = wb.active
                ws.title = "Games left to clean"
                all_games.sort();
                for row in zip(all_games):
                    ws.append(row)
                wb.save(filename='games left to clean.xlsx')
                print('Memory Written')

        def read_from_memory():
            wb = Workbook()
            print('Good')
            ws = wb.active
            ws.title = "Games to Clean"
            book = xlrd.open_workbook('games left to clean.xlsx')
            sheet = book.sheet_by_name('Games left to clean')
            all_games = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
            all_games = [item for sublist in all_games for item in sublist]
            print(len(all_games))
            bookemp = xlrd.open_workbook('masterlist.xlsx')
            sheet = bookemp.sheet_by_name('Employees')
            employee_list = [[sheet.cell_value(q, t) for t in range(sheet.ncols)] for q in range(sheet.nrows)]
            employee_list = [item for sublist in employee_list for item in sublist]
            num = (len(employee_list))
            num2 = (len(all_games))
            if num >= num2:
                print(
                    'Not enough games to distribute for this week, writing remaining games to worksheet, make note of them before reseting')
                for row in zip(all_games):
                    ws.append(row)
                ws.column_dimensions['A'].width = 50
                wb.save(filename='data.xlsx')
                return 'No games left'
            else:
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws['A1'] = 'Employee'
                ws['B1'] = 'Game to Clean'
                ws['C1'] = 'Manager Initial'
                ws['D1'] = 'Date'
                games_to_clean = []
                count = 0
                while count < num:
                    random.shuffle(all_games)
                    x = all_games.pop()
                    games_to_clean.append(x)
                    count = count + 1
                for row in zip(employee_list, games_to_clean):
                    ws.append(row)
                print(all_games)
                print(len(all_games))
                wb.save(filename='data.xlsx')
                print('DONE')
                return all_games

        def weeklygames(username):
            os.system('cls')
            while True:
                try:
                    os.system('cls')
                    print('1 takes you to manager menu')
                    print('2 assigns games to clean from Memory')
                    print('3 resets the games cleaned list back to original and assigns games afterwords (Reset)')
                    num = int((input('Enter a number: ')))
                    if num == 3:
                        all_games = overwrite()
                        writememory(all_games)
                        input("Games Cleaned Reset, press enter to continue")
                    if num == 2:
                        all_games = read_from_memory()
                        writememory(all_games)
                        input("Games Cleaned written from memory, press enter to continue")
                    if num == 1:
                        manager_options(username)
                except:
                    print('Invalid Input, please try again')
                    input('Press Enter to try again')
                    pass

        def manager_options(username):

            menu = ["1.Cleaning Tasks Currently in Progress", "2.Weekly Games Cleaning Tasks", "3.Exit"]
            while True:
                try:
                    os.system('cls')
                    print('\n'.join(str(p) for p in menu))

                    manager_select = int((input("Enter a number: ")))
                    if manager_select == 3:
                        input('You selected 3, going back to the login screen, press enter to continue')
                        main()
                    elif manager_select == 2:
                        input('Weekly Games Cleaning Tasks, press enter to continue')
                        weeklygames(username)
                    elif manager_select == 1:
                        input('Cleaning Tasks currently in progress, press enter to continue')
                        managerselection(username)
                except:
                    print('Invalid Input, please try again')
                    input('Press Enter to try again')
                    pass

        #############################
        def main():
            while True:
                try:
                    os.system('cls')
                    print('Welcome to the Nickelrama Cleaning Program')
                    print('These are the tasks that are currently in progress')
                    print('      ')
                    wb = load_workbook('Cleaning.xlsx')
                    ws = wb.get_sheet_by_name("Cleaning")
                    x = 'In progress'
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == x:
                                val = cell.value
                                c = cell.coordinate
                                offset3 = cell.offset(0, -1).value
                                datedata = cell.offset(0, -3).value
                                offset2 = cell.offset(0, -2).value
                                offset1 = cell.offset(0, -3).value
                                print('   ')
                                print("{0} | {1} | {2} | {3}".format(offset1, offset2, offset3, val))
                                print('    ')
                    Cody, username = login()
                    if Cody == 0:
                        print('LOL IM LOST')
                    if Cody == 1:
                        print('Hello Trainee')
                        print('       ')
                        traineeselections(username)
                    elif Cody == 2:
                        print('Hello Attendent')
                        print('       ')
                        attendantselections(username)
                    elif Cody == 3:
                        print('Hello Staff Attendent')
                        print('       ')
                        staffselections(username)
                    elif Cody == 4:
                        print('Senior Staff')
                        print('       ')
                        seniorselections(username)
                    elif username == 5:
                        manager_options(username)

                except TypeError:
                    print('Name not recognized, please try again.')
                    input('Press enter to continue')
                    pass

        ##############
        main()


app=Application()

mainloop()

